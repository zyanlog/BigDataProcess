#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

totalList = []
fcount = 0
for row in ws.iter_rows(min_row=2):
	midterm = row[2].value
	final = row[3].value
	homework = row[4].value
	attendance = row[5].value

	sum = midterm * 0.3 + final * 0.35 + homework * 0.34 + attendance

	row[6].value = sum
	if sum < 40:
		fcount += 1
	totalList.append(sum)

totalList.sort(reverse=True)

a = 0; ap = 0; b = 0; bp = 0; c = 0; cp = 0; #등급별 인원수

if fcount > len(totalList) - int(len(totalList) * 0.3):
	a = len(totalList) - fcount
	ap = int(a * 0.5)
elif fcount > len(totalList) - int(len(totalList) * 0.7):
	a = int(len(totalList) * 0.3)
	ap = int(a * 0.5)
	b = len(totalList) - a - fcount
	bp = int(b * 0.5)
else:
	a = int(len(totalList) * 0.3)
	ap = int(a * 0.5)
	b = int(len(totalList) * 0.7) - a
	bp = int(b * 0.5)
	c = len(totalList) - a - b - fcount
	cp = int(c * 0.5)

for row in ws.iter_rows(min_row=2):
	if row[6].value < 40:
		row[7].value = 'F'
	elif row[6].value >= totalList[ap - 1]:
		row[7].value = 'A+'
	elif row[6].value >= totalList[a - 1]:
		row[7].value = 'A0'
	elif row[6].value >= totalList[a + bp - 1] and bp != 0:
		row[7].value = 'B+'
	elif row[6].value >= totalList[a + b - 1] and b != 0:
		row[7].value = 'B0'
	elif row[6].value >= totalList[a + b + cp - 1] and cp != 0:
		row[7].value = 'C+'
	elif row[6].value >= totalList[a + b + cp + c - 1] and c != 0:
		row[7].value = 'C0'

wb.save('student.xlsx')
